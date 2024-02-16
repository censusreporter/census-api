from openpyxl.styles import Alignment, Font
from sqlalchemy import text
import logging
import openpyxl

logger = logging.getLogger('exporters')


def get_sql_config(session):
    """Return a tuple of strings: (host, port, user, password, database)"""
    bind = session.get_bind()
    return (bind.url.host,
            bind.url.port,
            bind.url.username,
            bind.url.password,
            bind.url.database)


def create_excel_download(session, data, table_metadata, valid_geo_ids, file_ident, out_filename, format):
    def excel_helper(sheet, table_id, table, option):
        """
        Create excel sheet.

        :param option: 'value' or 'percent'
        """

        # Write table id and title on the first row
        sheet['A1'] = table_id
        sheet['B1'] = table['title']

        sheet['A1'].font = Font(bold=True)
        sheet['B1'].font = Font(bold=True)

        header = []
        max_indent = 0
        # get column headers
        for column_id, column_info in table['columns'].items():
            column_name_utf8 = column_info['name'].encode('utf-8')
            indent = column_info['indent']

            header.append((column_name_utf8, indent))

            if indent > max_indent:
                max_indent = indent

        # Populate first column with headers
        for i, col_tuple in enumerate(header):
            header = col_tuple[0]
            current_row = i + 4  # 1-based index, account for geographic headers
            current_cell = sheet.cell(row=current_row, column=1)
            current_cell.value = header
            indent = col_tuple[1]
            if indent is None:
                logger.warn("Null indent for {} {}".format(table_id, header))
                indent = 0
            current_cell.alignment = Alignment(indent=indent, wrap_text=True)

        # Resize column width
        sheet.column_dimensions['A'].width = 50

        # this SQL echoed in OGR export but no geom so copying instead of factoring out
        # plus different binding when using SQLAlchemy
        result = session.execute(text(
            """SELECT full_geoid,display_name
                     FROM tiger2022.census_name_lookup
                     WHERE full_geoid IN :geoids
                     ORDER BY full_geoid"""),
            {'geoids': tuple(valid_geo_ids)}
        )

        geo_headers = []
        any_zero_denominators = False
        has_denominator_column = False
        for i, (geoid, name) in enumerate(result):
            geo_headers.append(name)
            col_values = []
            col_errors = []
            for table_id, table in table_metadata.items():
                table_estimates = data.get(geoid, {}).get(table_id, {}).get('estimate')
                table_errors = data.get(geoid, {}).get(table_id, {}).get('error')

                if option == 'value':
                    for column_id, column_info in table['columns'].items():
                        col_values.append(table_estimates.get(column_id, ''))
                        col_errors.append(table_errors.get(column_id, ''))
                elif option == 'percent':
                    denominator_column_id = table.get('denominator_column_id')
                    if denominator_column_id:
                        has_denominator_column = True
                        base_estimate = data[geoid][table_id]['estimate'][denominator_column_id]

                        for column_id, column_info in table['columns'].items():
                            column_estimate_value = table_estimates.get(column_id)
                            column_error_value = table_errors.get(column_id)
                            if base_estimate and column_estimate_value is not None and column_error_value is not None:
                                col_values.append(column_estimate_value / base_estimate)
                                col_errors.append(column_error_value / base_estimate)
                            else:
                                any_zero_denominators = True
                                col_values.append('*')
                                col_errors.append('')
                    else:
                        col_values.append('*')
                        col_errors.append('')

            for j, value in enumerate(col_values):
                col_num = (i + 1) * 2
                row_num = j + 4
                sheet.cell(row=row_num, column=col_num).value = value
                sheet.cell(row=row_num, column=col_num + 1).value = col_errors[j]
                if option == 'percent':
                    sheet.cell(row=row_num, column=col_num).number_format = '0.00%'
                    sheet.cell(row=row_num, column=col_num + 1).number_format = '0.00%'

        if option == 'percent' and (any_zero_denominators or not has_denominator_column):
            annotation_cell = sheet.cell(row=(row_num + 1), column=1)
            annotation_cell.font = Font(italic=True)
            if any_zero_denominators:
                annotation_cell.value = "* Base value of zero; no percentage available"
            elif not has_denominator_column:
                annotation_cell.value = "* Percentage values not appropriate for this table"
            else:
                annotation_cell.value = "* Unexpected error. Please contact Census Reporter at https://censusreporter.uservoice.com/ and let us know the page from where you downloaded this data."

        # Write geo headers
        for i in range(len(geo_headers)):
            current_col = (i + 1) * 2
            current_cell = sheet.cell(row=2, column=current_col)
            current_cell.value = geo_headers[i]
            current_cell.alignment = Alignment(horizontal='center')
            sheet.merge_cells(start_row=2, end_row=2, start_column=current_col, end_column=current_col + 1)
            sheet.cell(row=3, column=current_col).value = "Value"
            sheet.cell(row=3, column=current_col + 1).value = "Error"

    wb = openpyxl.workbook.Workbook()

    # For every table in table_metadata, make a two sheets (values and percentages)
    for i, (table_id, table) in enumerate(table_metadata.items()):
        sheet = wb.active
        sheet.title = table_id + ' Values'
        excel_helper(sheet, table_id, table, 'value')

        sheet_percents = wb.create_sheet(table_id + ' Percentages')
        excel_helper(sheet_percents, table_id, table, 'percent')

    wb.save(out_filename)


def create_ogr_download(session, data, table_metadata, valid_geo_ids, file_ident, out_filename, format):
    from osgeo import ogr
    from osgeo import osr
    format_info = supported_formats[format]
    driver_name = format_info['driver']
    ogr.UseExceptions()
    in_driver = ogr.GetDriverByName("PostgreSQL")
    host, port, user, password, database = get_sql_config(session)
    conn = in_driver.Open("PG: host=%s port=%s dbname=%s user=%s password=%s" % (host, port, database, user, password))

    if conn is None:
        raise Exception("Could not connect to database to generate download.")

    out_driver = ogr.GetDriverByName(driver_name)
    out_srs = osr.SpatialReference()
    out_srs.ImportFromEPSG(4326)
    out_data = out_driver.CreateDataSource(out_filename)
    # See http://gis.stackexchange.com/questions/53920/ogr-createlayer-returns-typeerror
    out_layer = out_data.CreateLayer(file_ident, srs=out_srs, geom_type=ogr.wkbMultiPolygon)
    out_layer.CreateField(ogr.FieldDefn('geoid', ogr.OFTString))
    out_layer.CreateField(ogr.FieldDefn('name', ogr.OFTString))
    for (table_id, table) in table_metadata.items():
        for column_id, column_info in table['columns'].items():
            if driver_name == "ESRI Shapefile":
                # Work around the Shapefile column name length limits
                out_layer.CreateField(ogr.FieldDefn(column_id, ogr.OFTReal))
                out_layer.CreateField(ogr.FieldDefn(column_id + "e", ogr.OFTReal))
            else:
                out_layer.CreateField(ogr.FieldDefn(column_id, ogr.OFTReal))
                out_layer.CreateField(ogr.FieldDefn(column_id + ", Error", ogr.OFTReal))

    # this SQL echoed in Excel export but no geom so copying instead of factoring out
    sql = """SELECT geom,full_geoid,display_name
             FROM tiger2022.census_name_lookup
             WHERE full_geoid IN (%s)
             ORDER BY full_geoid""" % ', '.join("'%s'" % g for g in valid_geo_ids)
    in_layer = conn.ExecuteSQL(sql)

    in_feat = in_layer.GetNextFeature()
    while in_feat is not None:
        out_feat = ogr.Feature(out_layer.GetLayerDefn())
        if format in ('shp', 'kml', 'geojson'):
            out_feat.SetGeometry(in_feat.GetGeometryRef())
        geoid = in_feat.GetField('full_geoid')
        out_feat.SetField('geoid', geoid)
        out_feat.SetField('name', in_feat.GetField('display_name'))
        for (table_id, table) in table_metadata.items():
            table_estimates = data[geoid][table_id]['estimate']
            table_errors = data[geoid][table_id]['error']
            for column_id, column_info in table['columns'].items():
                if column_id in table_estimates:
                    if format == 'shp':
                        # Work around the Shapefile column name length limits
                        estimate_col_name = column_id
                        error_col_name = column_id + "e"
                    else:
                        estimate_col_name = column_id
                        error_col_name = column_id + ", Error"

                    out_feat.SetField(estimate_col_name, table_estimates[column_id])
                    out_feat.SetField(error_col_name, table_errors[column_id])

        out_layer.CreateFeature(out_feat)
        in_feat.Destroy()
        in_feat = in_layer.GetNextFeature()
    out_data.Destroy()


supported_formats = {  # these should all have a 'function' with the right signature
    'shp': {"function": create_ogr_download, "driver": "ESRI Shapefile"},
    'kml': {"function": create_ogr_download, "driver": "KML"},
    'geojson': {"function": create_ogr_download, "driver": "GeoJSON"},
    'xlsx': {"function": create_excel_download, "driver": "XLSX"},
    'csv': {"function": create_ogr_download, "driver": "CSV"},
}
